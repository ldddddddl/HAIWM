import os
import sys
import argparse
from typing import Dict, List, Tuple, Optional

import numpy as np
from PIL import Image
from tqdm import tqdm
import openpyxl

try:
    from decord import VideoReader
    DECORD_AVAILABLE = True
except Exception:
    DECORD_AVAILABLE = False

TASK = {
    'v3': 'Put the block in another random position.',
    'v4': 'Pull the box closer and put the block into the box.'
}

def normalize_sheet_name(name: str) -> str:
    n = name.strip().lower()
    if n in {"position", "positions"}:
        return "positions"
    if n in {"velocity", "velocities", "velocites"}:
        return "velocities"
    if n in {"effort", "efforts", "effors"}:
        return "efforts"
    if n in {"sucker_action", "sucker_actions"}:
        return "sucker_actions"
    return n


def safe_as_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        return float(x)
    except Exception:
        return None


def read_excel_sheets(xlsx_path: str) -> Dict[str, Dict[str, np.ndarray]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: Dict[str, Dict[str, np.ndarray]] = {}
    for sheet in wb.worksheets:
        sname = normalize_sheet_name(sheet.title)
        if sname not in {"positions", "velocities", "efforts", "sucker_actions"}:
            continue
        first_row = list(next(sheet.iter_rows(values_only=True, max_row=1)))
        t_vals: List[float] = []
        valid_col_indices: List[int] = []
        for c_idx, v in enumerate(first_row):
            v_float = safe_as_float(v)
            if v_float is not None:
                t_vals.append(v_float)
                valid_col_indices.append(c_idx)
        if len(t_vals) == 0:
            continue
        t = np.asarray(t_vals, dtype=np.float64)
        data_rows: List[List[float]] = []
        for r in sheet.iter_rows(min_row=2, values_only=True):
            row_vals: List[float] = []
            for c_idx in valid_col_indices:
                v = r[c_idx] if c_idx < len(r) else None
                if sname == "sucker_actions":
                    if v is None:
                        row_vals.append(0.0)
                    elif isinstance(v, str):
                        x = v.strip().lower()
                        if x in {"off", "0", "false"}:
                            row_vals.append(0.0)
                        else:
                            row_vals.append(1.0)
                    else:
                        vv = safe_as_float(v)
                        row_vals.append(1.0 if vv is not None and vv > 0.5 else 0.0)
                else:
                    v_float = safe_as_float(v)
                    row_vals.append(0.0 if v_float is None else float(v_float))
            if len(row_vals) > 0:
                data_rows.append(row_vals)
        if len(data_rows) == 0:
            continue
        data = np.asarray(data_rows, dtype=np.float64)
        t_max = t[-1] if t[-1] > 0 else (t.max() if t.max() > 0 else 1.0)
        t_norm = t / t_max
        out[sname] = {"t": t_norm, "data": data}
    return out


def build_common_video_timeline(num_frames: int) -> np.ndarray:
    if num_frames <= 1:
        return np.zeros((num_frames,), dtype=np.float64)
    return np.linspace(0.0, 1.0, num_frames, endpoint=True)


def interpolate_channels(src_t: np.ndarray, src_data: np.ndarray, dst_t: np.ndarray, *, is_discrete: bool = False) -> np.ndarray:
    C, _ = src_data.shape
    T_dst = dst_t.shape[0]
    out = np.zeros((C, T_dst), dtype=np.float64)
    for c in range(C):
        y = src_data[c]
        if is_discrete:
            interp = np.interp(dst_t, src_t, y, left=y[0], right=y[-1])
            out[c] = np.rint(interp)
        else:
            out[c] = np.interp(dst_t, src_t, y, left=y[0], right=y[-1])
    return out


def read_video_frames(video_path: str, max_frames: Optional[int] = None) -> List[np.ndarray]:
    if not os.path.exists(video_path):
        raise FileNotFoundError(video_path)
    frames: List[np.ndarray] = []
    if DECORD_AVAILABLE:
        vr = VideoReader(video_path)
        n = len(vr)
        use_n = n if max_frames is None else min(n, max_frames)
        for i in range(use_n):
            img = vr[i].asnumpy()
            frames.append(img)
    else:
        import cv2
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            raise RuntimeError(f"无法打开视频: {video_path}")
        cnt = 0
        while True:
            ok, frame_bgr = cap.read()
            if not ok:
                break
            frame_rgb = frame_bgr[:, :, ::-1].copy()
            frames.append(frame_rgb)
            cnt += 1
            if max_frames is not None and cnt >= max_frames:
                break
        cap.release()
    return frames


def get_video_len_fps(video_path: str) -> Tuple[int, float]:
    if DECORD_AVAILABLE:
        vr = VideoReader(video_path)
        n = len(vr)
        try:
            fps = float(vr.get_avg_fps())
        except Exception:
            fps = 30.0
        return n, fps
    else:
        import cv2
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            raise RuntimeError(f"无法打开视频: {video_path}")
        n = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = float(cap.get(cv2.CAP_PROP_FPS) or 30.0)
        cap.release()
        return n, fps


def infer_video_paths(files: List[str]) -> Tuple[Optional[str], Optional[str]]:
    vids = [f for f in files if f.lower().endswith((".mp4", ".avi", ".mov", ".mkv"))]
    gripper = None
    side = None
    for v in vids:
        v_lower = v.lower()
        if "gripper" in v_lower and gripper is None:
            gripper = v
        elif "side" in v_lower and side is None:
            side = v
    if gripper is None and side is None and len(vids) >= 2:
        vids_sorted = sorted(vids)
        gripper, side = vids_sorted[0], vids_sorted[1]
    return gripper, side


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def hwc_to_chw_uint8(img: np.ndarray) -> np.ndarray:
    if img.dtype != np.uint8:
        if img.max() <= 1.0:
            img = (img * 255.0).astype(np.uint8)
        else:
            img = img.astype(np.uint8)
    return np.transpose(img, (2, 0, 1))


def resize_image_np(img_chw: np.ndarray, out_h: int, out_w: int) -> np.ndarray:
    c, h, w = img_chw.shape
    pil = Image.fromarray(np.transpose(img_chw, (1, 2, 0)))
    pil = pil.resize((out_w, out_h), Image.BICUBIC)
    arr = np.array(pil)
    if arr.ndim == 2:
        arr = np.repeat(arr[:, :, None], 3, axis=2)
    return np.transpose(arr, (2, 0, 1)).astype(np.uint8)


def peek_first_episode(input_root: str) -> Optional[str]:
    candidates = [os.path.join(input_root, d) for d in os.listdir(input_root) if os.path.isdir(os.path.join(input_root, d))]
    return sorted(candidates)[0] if candidates else None


def analyze_episode(episode_dir: str) -> Dict[str, int]:
    files = os.listdir(episode_dir)
    excel_candidates = [f for f in files if f.lower().endswith((".xlsx", ".xls"))]
    if len(excel_candidates) == 0:
        raise FileNotFoundError(f"未找到 Excel：{episode_dir}")
    excel_path = os.path.join(episode_dir, excel_candidates[0])
    sheets = read_excel_sheets(excel_path)
    if "positions" not in sheets:
        raise RuntimeError(f"Excel 缺少 positions sheet: {excel_path}")
    num_joints = sheets["positions"]["data"].shape[0]
    has_vel = 1 if "velocities" in sheets else 0
    has_eff = 1 if "efforts" in sheets else 0
    has_suck = 1 if "sucker_actions" in sheets and sheets["sucker_actions"]["data"].size > 0 else 0
    return {"num_joints": num_joints, "has_vel": has_vel, "has_eff": has_eff, "has_suck": has_suck}


def main():
    parser = argparse.ArgumentParser(description="Convert to LeRobotDataset format and optional push to hub.")
    parser.add_argument("--input_root", type=str, default=r"./datasets/v3", help="原始数据根目录。每个子目录为一个 episode")
    parser.add_argument("--output_root", type=str, default=r"./datasets/v3_lerobot", help="输出数据根目录")
    parser.add_argument("--image_size", type=int, default=256, help="输出图像边长，CHW = (3, S, S)")
    parser.add_argument("--overwrite", action="store_true",default=True, help="若输出目录存在则清空")
    parser.add_argument("--push_to_hub", action="store_true", default=False, help="是否推送到 Hugging Face Hub")
    parser.add_argument("--repo_id", type=str, default=None, help="Hugging Face repo_id，例如 username/dataset_name")
    parser.add_argument("--private", action="store_true", help="创建私有数据集仓库")
    parser.add_argument("--store_video_refs", action="store_true", default=True, help="以视频路径+帧索引形式保存，而不是每帧图像")
    parser.add_argument("--target_fps", type=int, default=30, help="目标帧率（下采样至该帧率），仅影响样本间隔")
    parser.add_argument("--no_prompt", action="store_true", help="不在每帧重复保存提示文本以减小体积")
    args = parser.parse_args()
    
    task_name = TASK[args.input_root.split('/')[-1]]
    # 延迟导入以避免环境缺失时报错
    try:
        from lerobot.datasets.lerobot_dataset import LeRobotDataset
    except Exception:
        print("请先安装并可导入 lerobot 库: pip install lerobot")
        raise

    input_root = args.input_root
    output_root = args.output_root
    if not os.path.isdir(input_root):
        raise NotADirectoryError(input_root)

    if os.path.exists(output_root) and args.overwrite:
        import shutil
        shutil.rmtree(output_root)
    # os.makedirs(output_root, exist_ok=True)

    first_ep = peek_first_episode(input_root)
    if first_ep is None:
        print("输入目录下未找到任何子目录。")
        sys.exit(1)
    info = analyze_episode(first_ep)
    num_joints = info["num_joints"]
    has_vel = info["has_vel"]
    has_eff = info["has_eff"]
    has_suck = info["has_suck"]

    # 创建数据集（参考 conver_example.py 的写入方式）
    if args.store_video_refs:
        features: Dict[str, Dict] = {
            # 用路径+索引引用视频帧，避免展开每帧图像
            "top_video_path": {
                "dtype": "string",
                "shape": (1,),
            },
            "top_frame_idx": {
                "dtype": "int32",
                "shape": (1,),
            },
            "wrist_video_path": {
                "dtype": "string",
                "shape": (1,),
            },
            "wrist_frame_idx": {
                "dtype": "int32",
                "shape": (1,),
            },
            "state": {
                "dtype": "float32",
                "shape": (num_joints + (1 if has_suck else 0),),
                "names": ["state"],
            },
            "action": {
                "dtype": "float32",
                "shape": (num_joints + (1 if has_suck else 0),),
                "names": ["action"],
            },
        }
    else:
        features: Dict[str, Dict] = {
            "top_image": {
                "dtype": "image",
                "shape": (3, args.image_size, args.image_size),
                "names": ["channel", "height", "width"],
            },
            "wrist_image": {
                "dtype": "image",
                "shape": (3, args.image_size, args.image_size),
                "names": ["channel", "height", "width"],
            },
            "state": {
                "dtype": "float32",
                "shape": (num_joints + (1 if has_suck else 0),),
                "names": ["state"],
            },
            "action": {
                "dtype": "float32",
                "shape": (num_joints + (1 if has_suck else 0),),
                "names": ["action"],
            },
        }
    if not args.no_prompt:
        features["prompt"] = {
            "dtype": "string",
            "shape": (1,),
            "names": ["instruction"],
        }
    if has_vel:
        features["velocities"] = {
            "dtype": "float32",
            "shape": (num_joints,),
            "names": ["velocities"],
        }
    if has_eff:
        features["efforts"] = {
            "dtype": "float32",
            "shape": (num_joints,),
            "names": ["efforts"],
        }
    if has_suck:
        features["sucker"] = {
            "dtype": "int32",
            "shape": (1,),
            "names": ["sucker"],
        }
        features["sucker_action"] = {
            "dtype": "int32",
            "shape": (1,),
            "names": ["sucker_action"],
        }

    dataset = LeRobotDataset.create(
        repo_id=args.repo_id,
        root=output_root,
        fps=args.target_fps,
        features=features,
        image_writer_threads=8,
        image_writer_processes=8,
    )

    episodes = [os.path.join(input_root, d) for d in os.listdir(input_root) if os.path.isdir(os.path.join(input_root, d))]
    episodes = sorted(episodes)

    for ep_dir in tqdm(episodes, desc="Processing episodes"):
        files = os.listdir(ep_dir)
        # Excel
        excel_candidates = [f for f in files if f.lower().endswith((".xlsx", ".xls"))]
        if len(excel_candidates) == 0:
            print(f"[跳过] 未找到 Excel: {ep_dir}")
            continue
        excel_path = os.path.join(ep_dir, excel_candidates[0])
        sheets = read_excel_sheets(excel_path)
        if "positions" not in sheets:
            print(f"[跳过] Excel 缺少 positions: {excel_path}")
            continue

        # 视频
        v_gripper_name, v_side_name = infer_video_paths(files)
        if v_gripper_name is None or v_side_name is None:
            print(f"[跳过] 未找到两个视频(gripper/side)：{ep_dir}")
            continue
        v_gripper = os.path.join(ep_dir, v_gripper_name)
        v_side = os.path.join(ep_dir, v_side_name)

        # 若保存为视频引用，避免解码所有帧，仅查询长度与 fps
        if args.store_video_refs:
            n_g, fps_g = get_video_len_fps(v_gripper)
            n_s, fps_s = get_video_len_fps(v_side)
            num_frames = min(n_g, n_s)
            # 计算步长以匹配目标 fps（约等于）
            stride_g = max(1, int(round((fps_g or args.target_fps) / args.target_fps)))
            stride_s = max(1, int(round((fps_s or args.target_fps) / args.target_fps)))
            stride = max(1, min(stride_g, stride_s))
        else:
            frames_gripper = read_video_frames(v_gripper)
            frames_side = read_video_frames(v_side)
            num_frames = min(len(frames_gripper), len(frames_side))
            stride = 1
        if num_frames < 2:
            print(f"[跳过] 帧数不足 2：{ep_dir}")
            continue

        # 时间轴
        t_video = build_common_video_timeline(num_frames)

        # 对齐通道
        pos_t = sheets["positions"]["t"]
        pos_data = sheets["positions"]["data"]
        pos_aligned = interpolate_channels(pos_t, pos_data, t_video, is_discrete=False)

        vel_aligned = None
        if "velocities" in sheets:
            vel_t = sheets["velocities"]["t"]
            vel_data = sheets["velocities"]["data"]
            vel_aligned = interpolate_channels(vel_t, vel_data, t_video, is_discrete=False)

        eff_aligned = None
        if "efforts" in sheets:
            eff_t = sheets["efforts"]["t"]
            eff_data = sheets["efforts"]["data"]
            eff_aligned = interpolate_channels(eff_t, eff_data, t_video, is_discrete=False)

        suck_aligned = None
        if "sucker_actions" in sheets and sheets["sucker_actions"]["data"].size > 0:
            suck_t = sheets["sucker_actions"]["t"]
            suck_data = sheets["sucker_actions"]["data"]
            suck_data_bin = suck_data[:1, :] if suck_data.ndim == 2 else suck_data.reshape(1, -1)
            suck_aligned = interpolate_channels(suck_t, suck_data_bin, t_video, is_discrete=True)[0]

        # 如需视频引用，准备输出视频副本相对路径
        if args.store_video_refs:
            import shutil
            ensure_dir(os.path.join(output_root, "videos"))
            ep_name = os.path.basename(os.path.normpath(ep_dir))
            side_ext = os.path.splitext(v_side)[1]
            grip_ext = os.path.splitext(v_gripper)[1]
            side_dst = os.path.join(output_root, "videos", f"{ep_name}_side{side_ext}")
            grip_dst = os.path.join(output_root, "videos", f"{ep_name}_gripper{grip_ext}")
            if not os.path.exists(side_dst):
                shutil.copy2(v_side, side_dst)
            if not os.path.exists(grip_dst):
                shutil.copy2(v_gripper, grip_dst)
            side_rel = os.path.relpath(side_dst, output_root).replace("\\", "/")
            grip_rel = os.path.relpath(grip_dst, output_root).replace("\\", "/")

        # 写入帧：状态用 t，动作用 t+1（支持按 stride 下采样）
        last_index = num_frames - 1
        idx = 0
        while idx < last_index:
            state_t = pos_aligned[:, idx].astype(np.float32)
            action_t1 = pos_aligned[:, idx + 1].astype(np.float32)

            current_sucker_int: Optional[int] = None
            next_sucker_int: Optional[int] = None
            if suck_aligned is not None:
                current_sucker_int = int(np.rint(suck_aligned[idx]))
                next_sucker_int = int(np.rint(suck_aligned[idx + 1]))
                # 将 sucker 合并到 state 与 action 的最后一维
                state_t = np.concatenate(
                    [state_t, np.array([float(current_sucker_int)], dtype=np.float32)], axis=0
                )
                action_t1 = np.concatenate(
                    [action_t1, np.array([float(next_sucker_int)], dtype=np.float32)], axis=0
                )

            if args.store_video_refs:
                frame = {
                    "top_video_path": side_rel,
                    "top_frame_idx": np.array([idx], dtype=np.int32),
                    "wrist_video_path": grip_rel,
                    "wrist_frame_idx": np.array([idx], dtype=np.int32),
                    "state": state_t,
                    "action": action_t1,
                }
            else:
                g_img = frames_gripper[idx]
                s_img = frames_side[idx]
                g_chw = hwc_to_chw_uint8(g_img)
                s_chw = hwc_to_chw_uint8(s_img)
                if g_chw.shape[1] != args.image_size or g_chw.shape[2] != args.image_size:
                    g_chw = resize_image_np(g_chw, args.image_size, args.image_size)
                if s_chw.shape[1] != args.image_size or s_chw.shape[2] != args.image_size:
                    s_chw = resize_image_np(s_chw, args.image_size, args.image_size)
                frame = {
                    "top_image": s_chw,
                    "wrist_image": g_chw,
                    "state": state_t,
                    "action": action_t1,
                }

            if not args.no_prompt:
                frame["prompt"] = task_name
            if vel_aligned is not None:
                frame["velocities"] = vel_aligned[:, idx].astype(np.float32)
            if eff_aligned is not None:
                frame["efforts"] = eff_aligned[:, idx].astype(np.float32)
            if suck_aligned is not None:
                frame["sucker"] = np.array([current_sucker_int], dtype=np.int32)
                frame["sucker_action"] = np.array([next_sucker_int], dtype=np.int32)

            if args.no_prompt:
                dataset.add_frame(frame)
            else:
                dataset.add_frame(frame, task=task_name)

            idx += stride

        dataset.save_episode()

    if args.push_to_hub:
        if not args.repo_id:
            print("--push_to_hub 需同时提供 --repo_id，例如 username/dataset_name")
        else:
            dataset.push_to_hub(
                tags=["robotics", "lerobot"],
                private=args.private,
                push_videos=False,
            )

    print(f"完成。输出目录: {output_root}")


if __name__ == "__main__":
    main()


