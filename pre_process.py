
import decord
from decord import VideoReader
import torch
import os
import openpyxl
# from publisher.jetmax_kinematics import joints2angle
import json
import numpy as np
from sys import getsizeof as getsize
import torch.nn.functional as F
import math
import torch.nn as nn
from PIL import Image
from matplotlib import pyplot as plt

# test
AngleRotateRange_joint2angle = -60, 60  # joint 1
AngleLeftRange_joint2angle = -40, 40    # joint 2
AngleRightRange_joint2angle = 0, 90     # joint 3

DATA_FOLDER = 'v3_30fps'
PER_IMAGES_WITH_SIGNAL_NUM = 20
PAST_IMG_NUM = 5
FUTURE_IMG_NUM = 5
BATCH_SIZE = 8
EPISODE_TO_BATCH_NUM = 1
IS_SPIKE = False
# Pass to model
# model(video_data)
class DecordVideo:
    def __init__(self) -> None:
        decord.bridge.set_bridge('torch')  # 将解码桥接至 torch 张量
        self.depress_list = []
    def process(self, path:str):
        try:
        # if True:
            vr = VideoReader(path)
            num_frames = len(vr)
            fps = vr.get_avg_fps()
            # 获取以 1Hz 采样的帧索引（每秒取一帧）
            samping_num = BATCH_SIZE*EPISODE_TO_BATCH_NUM*(PAST_IMG_NUM + FUTURE_IMG_NUM)
            samping_interval = num_frames/samping_num
            frame_indices = np.arange(0, num_frames, samping_interval)  # 每隔 fps 个帧采样一个帧
            video_data = vr.get_batch(frame_indices)  # 获取指定帧

            resize_d_tensor = F.interpolate(video_data.transpose(1, -1), size=(112, 112), mode='bicubic', align_corners=False).transpose(1, -1)
            # self.depress_list.append(resize_d_tensor)
            resize_d_tensor = resize_d_tensor.float() / 255.0  # 归一化
            return resize_d_tensor, True
        except:
            print(path)
            return path, False

class ReadFile:
    def __init__(self) -> None:
        parent_folder = os.path.dirname(os.path.abspath(__file__))
        self.folder_path = os.path.join(parent_folder, 'datasets', DATA_FOLDER) # dataset folder
        self.path_dict = {}


    def readfile(self):
        self.folder_list = os.listdir(self.folder_path)
        for f_cnt, f in enumerate(self.folder_list):
            path_ = os.path.join(self.folder_path, f)
            data_path = os.listdir(path_)
            self.path_dict[str(f_cnt)] = [os.path.join(path_, dp) for dp in data_path]
        return self.path_dict

class LIFNeuron(nn.Module):
    def __init__(self, threshold=0.5):
        super(LIFNeuron, self).__init__()
        self.threshold = threshold
        self.mem = None

    def forward(self, input_):
        if self.mem is None:
            self.mem = torch.zeros(input_.shape).to(input_.device)  # 初始化膜电位

        self.mem += input_  # 累积输入
        spikes = (self.mem > self.threshold).float()  # 发放脉冲
        self.mem -= self.threshold * spikes  # 重置膜电位
        return spikes

def split_data(data, n, m):
    # 按每组 n 个数据进行分组
    data_ = [list(col) for col in data]
    if len(data_) > n*m:
        excess_cnt = len(data_) - n*m
        data_ = data_[excess_cnt // 2 : len(data_) - excess_cnt // 2] # 掐头去尾
    groups = [data_[i * n:(i + 1) * n] for i in range(m)]
    
    return groups

def align_data(workbook, gripper_vr, side_vr, info, data_features, data_format='joints'):
    lif_spike = LIFNeuron()
    episode = {}
    positions_sheet = workbook['positions']
    first_row = list(next(positions_sheet.iter_rows(values_only=True, max_row=1)))
    divi_timestamp = [i/first_row[-1] for i in first_row]
    pos_cols = positions_sheet.iter_cols(values_only=True)
    sucker_actions_sheet = workbook['sucker_actions']
    # 将处理后的数据写入 sheet 的第一行
    for col, value in enumerate(divi_timestamp, start=1):  # 列索引从 1 开始
        positions_sheet.cell(row=1, column=col, value=value)
        sucker_actions_sheet.cell(row=1, column=col, value=value)
    sucker_cols = sucker_actions_sheet.iter_cols(values_only=True)
    max_col = positions_sheet.max_column
    grip_frames, h, w, c = gripper_vr.shape    
    side_frames, h, w, c = side_vr.shape    
    frames = grip_frames if side_frames > grip_frames else side_frames
    # per_frame_with_signal = max_col // frames
    per_frame_with_signal = PER_IMAGES_WITH_SIGNAL_NUM
    # print(f'Frames:{per_frame_with_signal}')
    # return None, None
    split_pos = split_data(pos_cols, per_frame_with_signal, frames)
    split_sucker = split_data(sucker_cols, per_frame_with_signal, frames)
    ######
    # if max_col > frames:
    #     # 计算要抽取的帧的索引
    #     indices = np.floor(np.linspace(0, max_col, frames, endpoint=False)).astype(int)
    #     pos_cols_list = list(pos_cols)
    #     sucker_cols_list = list(sucker_cols)
    #     selected_pos_cols = [pos_cols_list[i] for i in indices]
    #     selected_sucker_cols = [sucker_cols_list[i] for i in indices]
    #     selected_gripper_frames = gripper_vr
    #     selected_side_frames = side_vr
    #     data_length = frames

    # elif frames > max_col:
    #     indices = torch.tensor(np.floor(np.linspace(0, frames, max_col, endpoint=False)).astype(int))
    #     selected_gripper_frames = torch.index_select(gripper_vr, 0, indices)
    #     selected_side_frames = torch.index_select(side_vr, 0, indices)
    #     selected_pos_cols = list(pos_cols)
    #     selected_sucker_cols = list(sucker_cols)
    #     data_length = max_col

    # else:
    #     selected_pos_cols = list(pos_cols)
    #     selected_sucker_cols = list(sucker_cols)
    selected_gripper_frames = gripper_vr
    selected_side_frames = side_vr
    data_length = len(split_pos)
    

    for cnt, col in enumerate(zip(split_pos, split_sucker)):
        
        # timestamp = col[0][0][0]
        joint_pos_datas = col[0]
        sucker_action = col[1]
        for s_cnt, s in enumerate(sucker_action):
            if s[1] == 'off':
                sucker_action[s_cnt][1] = 0
            else:
                sucker_action[s_cnt][1] = 1
        if data_format == 'rpy':
            angle = [joints2angle(j) for j in joint_pos_datas] 
        else:
            angle = joint_pos_datas # no snn
            # tens_joint_pos_datas = torch.tensor(joint_pos_datas, dtype=torch.float32)
            # angle = lif_spike(tens_joint_pos_datas)

        action_dict = {
            # "motor_angle":{
            # "motor_1":angle[0],
            # "motor_2":angle[1],
            # "motor_3":angle[2]
            # },
            "joints_position":angle
            ,
            "sucker_actions":sucker_action
        }   
        if IS_SPIKE:  
            image_grp = lif_spike(selected_gripper_frames[cnt, ...])
            image_side = lif_spike(selected_side_frames[cnt, ...])
        else:
            image_grp = selected_gripper_frames[cnt, ...]
            image_side = selected_side_frames[cnt, ...]
                
        # plt.imshow(spike_image_side)
        # plt.show()
        obs_dict = {
            # "gripper_video":selected_gripper_frames[cnt, :, :, :],
            # "side_video":selected_side_frames[cnt, :, :, :]
            "gripper_video":image_grp,
            "side_video":image_side
        }
        episode[cnt] = {
            "action":action_dict,
            "observation":obs_dict
        }
        # print(col)
    # velocities_sheet = workbook['velocities']
    # efforts_sheet = workbook['efforts']
    return episode, data_length

# 将字典保存为 .npz 文件
def save_dict_to_npz(dictionary, filename):
    # 将所有内容扁平化
    def flatten_dict(d, parent_key=''):
        items = []
        for k, v in d.items():
            new_key = parent_key + '.' + str(k) if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key).items())
            else:
                items.append((new_key, v))
        return dict(items)

    flat_data = flatten_dict(dictionary)
    np.savez_compressed(filename, **flat_data)


AngleRotateRange = 0, 240
AngleLeftRange = 0, 180
AngleRightRange = -20, 160

def joints2angle(joint_angle):
    """
    Jetmax逆向运动学，从关节角度反推主动角度
    @param joint_angle: 包含9个关节角度的列表
    @return: 原始的三个主动角度 [rotate, angle left, angle right]，若无效返回None
    """
    if len(joint_angle) > 9:
        joint_angle = joint_angle[1:]
        # rospy.logerr("Invalid joint angles list! Expected 9 elements.")
    else:
        return None

    joint_angle = [(j*180.0) / math.pi for j in joint_angle]

    # 提取主动关节的角度
    joint1 = joint_angle[0]   # 对应rotate的关节
    joint2 = joint_angle[1]   # 对应angle left的关节
    joint6 = joint_angle[5]   # 对应angle right的关节

    # 逆向计算原始参数
    rotate = joint1 + 90
    angle_left = 90 - joint2
    angle_right = -joint6

    # 检查角度范围是否合法
    if (AngleRotateRange[0] <= rotate <= AngleRotateRange[1] and
        AngleLeftRange[0] <= angle_left <= AngleLeftRange[1] and
        AngleRightRange[0] <= angle_right <= AngleRightRange[1]):
        return [rotate, angle_left, angle_right]
    else:
        return None

if __name__ == "__main__":
    read_file = ReadFile()
    decorder = DecordVideo()
    path_dict = read_file.readfile()
    decord.bridge.set_bridge('torch')  # 将解码桥接至 torch 张量
    # with open('./data_features.json', "r") as f:
    #     data_features = json.load(f)
    data_features = {}
    bad_data = []

    for p_cnt, p in enumerate(path_dict):

        data_features = {}
        file_list = path_dict[p]
        print(p)
        # print(file_list)
        info = ''
        try:
            for fl in file_list:
                if fl.endswith("xlsx"):
                    workbook = openpyxl.load_workbook(fl)
                elif fl.endswith('avi'):
                    side_is_successful = True
                    grp_is_successful = True
                    if "gripper" in fl:
                        gripper_vr, grp_is_successful = decorder.process(fl)    
                    elif "side" in fl:
                        side_vr, side_is_successful = decorder.process(fl)
                elif fl.endswith('txt'):
                    with open(fl, "r", encoding="utf-8") as  f:
                        info = f.readlines()
                        info_temp = [i.replace(" ", "").replace("\n", "").split(":") for i in info]
                        info_dict = dict(i for i in info_temp[1:-1])

            # pass
            if not grp_is_successful:
                bad_data.append(gripper_vr)
                continue
            elif not side_is_successful:
                bad_data.append(side_vr)
                continue
            episode, data_length = align_data(workbook, gripper_vr, side_vr, info, data_features)    
                # continue
            print(bad_data)
            if info == '':
                data_features = {
                    # **info_dict, 
                    "video_frames_and_data_length":data_length,
                    "datas":{**episode},
                    }
            else:
                data_features = {
                    **info_dict, 
                    "video_frames_and_data_length":data_length,
                    "datas":{**episode},
                    }
            
            if not os.path.exists(f"datasets/{DATA_FOLDER}_npz"):
                os.mkdir(f"datasets/{DATA_FOLDER}_npz")
            save_dict_to_npz(data_features, f"datasets/{DATA_FOLDER}_npz/{DATA_FOLDER}-past_num_{PAST_IMG_NUM}-future_num_{FUTURE_IMG_NUM}-batchsize_{BATCH_SIZE}-per_episode_batch_num{EPISODE_TO_BATCH_NUM}-{p_cnt}.npz")
        
        except Exception as e:
            print(e)
            bad_data.append(fl)
            continue
        # print(getsize(data_features), data_length)
        # # 加载 .npy 文件并还原字典
        # loaded_dict = np.load('my_dict.npy', allow_pickle=True).item()
'''


'obj_pos_x':'93.9856'
'obj_pos_y':'-69.5899'
'obj_pos_z':'61.0000'
'obj_pos_r':'31.9000'
'gripper_pos_offset_x':'0.0000'
'gripper_pos_offset_y':'0.0000'
'gripper_pos_offset_z':'0.0000'
'gripper_pos_offset_r':'0.0000'
'putdown_pos_x':'-58'
'putdown_pos_y':'-178'
'putdown_pos_z':'61'
'obj_color':'Tertiary'
'model_name':'random_box0'
'is_fixed_tar_pos':'True'
'is_fixed_gen_pos':'True'
'is_sucessed':'True'
'other':''
'video_frames_and_data_length':149
'datas':{
    1:{
        actions:{
            joint_position:List
            sucker_actions:List
        }
        observation:{
            gripper_video:Tensor
            side_video:Tensor
        }
    }
    2:{
        ···
    }
    ···
    
}
'''